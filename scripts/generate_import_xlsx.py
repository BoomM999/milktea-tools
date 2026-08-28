# -*- coding: utf-8 -*-
"""
generate_import_xlsx.py

把"前场配料"PDF 转换为导入配方 Excel 的端到端脚本。

输入：
  --pdf <PDF路径>      必填
  --out <输出xlsx>      可选，默认写到桌面 <PDF文件名>-导入配方.xlsx
  --version <版本号>     可选，默认"全国北京路通用"
  --rows '<JSON列表>'    可选（高级），跳过 PDF 解析，直接用预解析好的 rows 写入。
                        每项 dict 包含 version/drink/cup/temp/sugar/material/dose 字段。

输出：
  成功：在 out 路径写出 .xlsx，并把解析过程打印到 stdout。
  失败：抛异常并打印结构化错误。

约束：
  - 不修改源 PDF，不写源 Excel。
  - 写入格式与模板完全一致（9 列 A-I，表头 row 1）。
  - 用量为字符串数字。
  - "不额外加糖" 组合不输出 蔗糖 行。
"""
import argparse
import json
import os
import re
import sys
from typing import List, Dict, Any

# 依赖：pymupdf（fitz 已 deprecated，但 1.x 仍可用），openpyxl
try:
    import fitz  # type: ignore
    PYMUPDF_OK = True
except Exception:
    try:
        import pymupdf as fitz  # type: ignore
        PYMUPDF_OK = True
    except Exception:
        PYMUPDF_OK = False

try:
    from openpyxl import Workbook
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


# ============ 常量（与模板实测一致）============
HEADERS = [
    "配方版本号（version）",
    "饮品名称（drinkName）",
    "杯型（cup size）",
    "做法（addtional）",
    "温度（ice level）",
    "糖度（sugar level）",
    "原料名称（materialName）",
    "原料用量（materialDosage）",
    "原料排序（materialOutSort）",
]

TEMPERATURE_ORDER = ["正常冰", "少冰", "去冰", "常温", "热"]
TEMPERATURE_ALIASES = {
    "全温度": TEMPERATURE_ORDER,
    "冰/少冰/去冰": ["正常冰", "少冰", "去冰"],
    "冰/少冰": ["正常冰", "少冰"],
    "冰": ["正常冰"],
    "冰沙": ["正常冰"],  # 2026-08-13 用户规则：冰沙选项一律改为正常冰
    "常温": ["常温"],
    "热": ["热"],
}

SUGAR_ORDER = ["标准糖度", "七分糖", "五分糖", "三分糖", "不额外加糖"]
MATERIAL_NAME_NORMALIZE = {
    "蔗糖糖浆": "蔗糖",
    "凤梨果酱": "凤梨果浆",
}

# 默认不写入 Excel 的非机器出料（黄金冻、罐头、果粒等）
NON_MACHINE_MATERIALS = {
    "黄金冻",
    "红西柚颗粒罐头",
}


# ============ PDF 解析（启发式，肉眼复核必做）============
def parse_pdf(pdf_path: str) -> Dict[str, Any]:
    """启发式解析 PDF，返回结构化数据。

    返回：
      {
        "version": str,
        "drinks": [
          {
            "name": "凤梨冰奶",
            "cup": "大杯",
            "temps": ["正常冰", ...],          # 该饮品允许的温度
            "sugars": [                        # 每档糖度对应的原料列表
              {"sugar": "标准糖度", "ingredients": [{"name":"蔗糖","dose":"30"}, ...]},
              {"sugar": "七分糖",   "ingredients": [{"name":"蔗糖","dose":"20"}]},
              ...
              {"sugar": "不额外加糖", "ingredients": [{"name":"凤梨果浆","dose":"25"}, ...]},
            ]
          },
          ...
        ],
        "rendered_pages": ["<png路径>", ...]   # 供肉眼复核
      }
    """
    if not PYMUPDF_OK:
        raise RuntimeError("缺少 pymupdf/fitz 依赖")
    doc = fitz.open(pdf_path)
    rendered = []
    outdir = os.path.join(os.path.dirname(pdf_path) or ".", "_pdf_preview")
    os.makedirs(outdir, exist_ok=True)
    for i in range(len(doc)):
        pix = doc[i].get_pixmap(matrix=fitz.Matrix(2, 2))
        fp = os.path.join(outdir, f"page{i+1}.png")
        pix.save(fp)
        rendered.append(fp)
        print(f"[preview] 渲染第{i+1}页: {fp}")

    full_text = "\n".join((p.get_text() or "") for p in doc)

    # 启发式：按饮品名切片（中文 2-12 字，含"奶/茶/露/椰/冰"等关键字）
    drink_keywords = ["冰奶", "奶茶", "甘露", "椰椰", "茶", "奶", "露"]
    candidate = re.findall(r"[\u4e00-\u9fff（）()]{2,12}", full_text)
    drink_names = []
    for w in candidate:
        if any(k in w for k in drink_keywords) and w not in drink_names:
            drink_names.append(w)
    print(f"[hint] 候选饮品名: {drink_names}")

    # 由于 PDF 表格的文本顺序经常被打乱，本脚本只给出占位结构；
    # 真实解析需要 LLM 配合渲染图人工核对后填入。
    return {
        "version": "全国北京路通用",
        "drinks": [],
        "rendered_pages": rendered,
        "raw_text": full_text,
        "candidate_names": drink_names,
    }


# ============ 结构化数据 → Excel ============
def expand_to_rows(struct: Dict[str, Any]) -> List[List[str]]:
    """把结构化数据展开成 Excel 行（不含表头）。"""
    rows: List[List[str]] = []
    version = struct.get("version") or "全国北京路通用"
    cup_default = struct.get("cup_default") or "大杯"

    for drink in struct["drinks"]:
        name = drink["name"]
        cup = drink.get("cup") or cup_default
        temps = drink.get("temps") or TEMPERATURE_ORDER
        sugars = drink.get("sugars") or []

        for temp in temps:
            for sug in sugars:
                sugar_label = sug["sugar"]
                ingredients = sug.get("ingredients", [])
                if sugar_label == "不额外加糖":
                    ingredients = [i for i in ingredients if i["name"] != "蔗糖"]
                if not ingredients:
                    continue
                for ing in ingredients:
                    if ing["name"] in NON_MACHINE_MATERIALS:
                        continue
                    rows.append([
                        version,        # A
                        name,           # B
                        cup,            # C
                        "",             # D 做法
                        temp,           # E
                        sugar_label,    # F
                        ing["name"],    # G
                        str(int(ing["dose"])),  # H
                        "",             # I 原料排序
                    ])
    return rows


def write_xlsx(out_path: str, rows: List[List[str]]) -> None:
    if not OPENPYXL_OK:
        raise RuntimeError("缺少 openpyxl 依赖")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    # 清理多余 Sheet2（openpyxl 默认只创建一个工作表，这里确保）
    while len(wb.sheetnames) > 1:
        del wb[wb.sheetnames[-1]]
    wb.save(out_path)


# ============ CLI ============
def main():
    ap = argparse.ArgumentParser(description="前场配料 PDF → 导入配方 Excel")
    ap.add_argument("--pdf", help="PDF 路径")
    ap.add_argument("--out", help="输出 xlsx 路径（默认写到桌面）")
    ap.add_argument("--version", default="全国北京路通用", help="配方版本号")
    ap.add_argument("--rows", help="跳过 PDF 解析，直接传入 JSON 行（含 version/drink/cup/temp/sugar/material/dose）")
    args = ap.parse_args()

    if not args.rows and not args.pdf:
        ap.error("必须提供 --pdf 或 --rows")

    if args.rows:
        raw = json.loads(args.rows)
        # raw 是 [{version, drink, cup, temp, sugar, material, dose}, ...]
        struct: Dict[str, Any] = {"version": args.version, "drinks": []}
        from collections import OrderedDict
        grouped: "OrderedDict[str, Dict]" = OrderedDict()
        for it in raw:
            key = it["drink"]
            d = grouped.setdefault(key, {"name": key, "cup": it.get("cup", "大杯"),
                                          "temps": [], "sugars": []})
            if it["temp"] not in d["temps"]:
                d["temps"].append(it["temp"])
            sug = next((s for s in d["sugars"] if s["sugar"] == it["sugar"]), None)
            if not sug:
                sug = {"sugar": it["sugar"], "ingredients": []}
                d["sugars"].append(sug)
            sug["ingredients"].append({"name": it["material"], "dose": it["dose"]})
        struct["drinks"] = list(grouped.values())
    else:
        parsed = parse_pdf(args.pdf)
        print("[hint] PDF 已解析为占位结构，请用 LLM 配合渲染图核对并填充 'drinks' 字段。")
        print(json.dumps({k: v for k, v in parsed.items() if k != "raw_text"}, ensure_ascii=False, indent=2))
        sys.exit(2)  # 需要人工核对，不直接写文件

    # 排序：温度/糖度按固定顺序
    t_rank = {t: i for i, t in enumerate(TEMPERATURE_ORDER)}
    s_rank = {s: i for i, s in enumerate(SUGAR_ORDER)}
    for drink in struct["drinks"]:
        drink["temps"] = sorted(set(drink["temps"]), key=lambda x: t_rank.get(x, 99))
        drink["sugars"] = sorted(drink["sugars"], key=lambda x: s_rank.get(x["sugar"], 99))

    rows = expand_to_rows(struct)

    out_path = args.out
    if not out_path:
        if not args.pdf:
            raise RuntimeError("无 --out 也无 --pdf，无法决定输出路径")
        out_dir = os.path.join(os.path.expanduser("~"), "Desktop", "openclaw")
        os.makedirs(out_dir, exist_ok=True)
        base = os.path.splitext(os.path.basename(args.pdf))[0]
        out_path = os.path.join(out_dir, f"{base}-导入配方.xlsx")

    write_xlsx(out_path, rows)
    print(f"[ok] 写入 {len(rows)} 行到 {out_path}")


if __name__ == "__main__":
    main()