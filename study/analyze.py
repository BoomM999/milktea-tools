import openpyxl
import os

base = os.path.dirname(os.path.abspath(__file__))
f1 = os.path.join(base, "标签简化_推荐糖配料表.xlsx")
f2 = os.path.join(base, "制作密码导出.xlsx")

def dump(path, label, max_rows=60, max_cols=40):
    print("="*80)
    print(f"FILE: {label}  ({os.path.basename(path)})")
    print("="*80)
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        print(f"\n--- SHEET: {ws.title}  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column} ---")
        rows = list(ws.iter_rows(values_only=True))
        for i, r in enumerate(rows):
            if i >= max_rows:
                print(f"... (truncated, total {len(rows)} rows)")
                break
            # trim trailing None
            r2 = list(r)
            while r2 and r2[-1] is None:
                r2.pop()
            print(f"[{i}] " + " | ".join("" if c is None else str(c) for c in r2))

dump(f1, "标签简化_推荐糖配料表")
dump(f2, "制作密码导出")
